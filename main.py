import openpyxl
from threading import Thread
from tqdm import tqdm
from validate_email import validate_email

def validate_emails(filename, start_row, end_row):
    # Load the workbook
    wb = openpyxl.load_workbook(filename)
    
    # Select the first worksheet
    ws = wb[wb.sheetnames[0]]
    
    # Open the files for writing
    with open("North_valid_emails.txt", "a") as valid_file, open("North_invalid_emails.txt", "a") as invalid_file:
        # Iterate through the rows in the worksheet
        for i, row in enumerate(ws.rows):
            # Skip rows that are not in the specified range
            if i < start_row or i >= end_row:
                continue
            
            # Get the name and email address from the first and second columns
            name = row[1].value
            email = row[2].value
            
            # Validate the email address
            is_valid = validate_email(email_address=email,
                                      check_format=True,
                                      check_blacklist=True,
                                      check_dns=True,
                                      dns_timeout=10,
                                      check_smtp=True,
                                      smtp_timeout=10,
                                      smtp_helo_host='my.host.name',
                                      smtp_from_address='admin@emancipation.co.in',
                                      smtp_skip_tls=False,
                                      smtp_tls_context=None,
                                      smtp_debug=False)
            if is_valid:
                # Write the name and email to the valid file
                valid_file.write(f"{name},{email}\n")
            else:
                # Write the email to the invalid file
                invalid_file.write(f"{email}\n")
            
            # Update the progress file
            with open("progress.txt", "w") as progress_file:
                progress_file.write(str(i))

def main():
    # Load the workbook
    wb = openpyxl.load_workbook("North.xlsx")
    
    # Select the first worksheet
    ws = wb[wb.sheetnames[0]]
    
    # Get the total number of rows in the worksheet
    total_rows = ws.max_row
    
    # Divide the rows into chunks of 1000
    chunk_size = 1000
    num_chunks = total_rows // chunk_size
    if total_rows % chunk_size > 0:
        num_chunks += 1
    
    # Create a progress bar
    pbar = tqdm(total=total_rows)
    
    # Create a list of threads
    threads = []
    
    # Start the threads
    for i in range(num_chunks):
        start_row = i * chunk_size
        end_row = start_row + chunk_size
        t = Thread(target=validate_emails, args=("North.xlsx", start_row, end_row))
        t.start()
        threads.append(t)
    
    # Wait for all threads to complete
    for t in threads:
        t.join()
    
    # Close the progress bar
    pbar.close()
    

if __name__ == "__main__":
    main()
